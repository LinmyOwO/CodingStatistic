from openpyxl import Workbook
from DataHandler import DataHandler

data_handler = DataHandler()


class TxtPrinter:
    ROW_LENGTH = 20

    def print(self, data, filename='statistic.txt'):

        with open(filename, 'w', encoding="utf-8") as file:
            count = 0
            number_string = ""
            time_string = ""

            for key, value in data.items():
                number_string += f" {key:6}|"
                time_string += f" {value:6.2f}|"
                count += 1

                if count == self.ROW_LENGTH:
                    file.write(number_string + '\n')
                    file.write('-' * len(number_string) + '\n')
                    file.write(time_string + '\n')
                    file.write('\n')

                    count = 0

            file.write(number_string + '\n')
            file.write('-' * len(number_string) + '\n')
            file.write(time_string + '\n')


class ExcelPrinter:
    def print(self, data, filename='statistic.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.title = "3-значные числа"

        numbers = list(data.keys())
        # numbers = data_handler.get_array_with_distinct_elem_length(numbers)
        numbers = sorted(numbers)

        ws.cell(1, 1, "Число")
        ws.cell(1, 2, "Частота")
        ws.cell(1, 3, "Лучшее время")
        ws.cell(1, 4, "Среднее время")
        ws.cell(1, 5, "Последний результат")

        for count, number in enumerate(numbers):
            num_stat = data[number]
            average_time = data_handler.find_average(num_stat['times'])

            ws.cell(count + 2, 1, number)
            ws.cell(count + 2, 2, num_stat['frequency'])
            ws.cell(count + 2, 3, num_stat['best_time'])
            ws.cell(count + 2, 4, average_time)
            ws.cell(count + 2, 5, num_stat['times'][-1])

        wb.save(filename)
